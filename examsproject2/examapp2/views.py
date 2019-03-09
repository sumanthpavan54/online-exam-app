from django.shortcuts import render, redirect
from django.contrib.auth.models import User, Group
from examapp2.forms import SignUpForm
from examapp2.forms import UserProfileInfoForm
from examapp2.forms import TotalScoreForm
from examapp2 import forms
from django.http import HttpResponseRedirect
from django.http import HttpResponse
from examapp2.models import TotalScore
from django.contrib.auth import authenticate, login, logout
import requests
from bs4 import BeautifulSoup
import openpyxl
# Create your views here.
def base1(request):
    return render(request,"examapp2/base1.html")
def signup_view(request):
    # form1=SignUpForm()
    # form2=UserProfileInfoForm()

    if request.method=='POST':
        form1=SignUpForm(data = request.POST)
        form2=UserProfileInfoForm(data = request.POST)
        if form1.is_valid() and form2.is_valid():
            user=form1.save()
            user.set_password(user.password)
            user.save()
            profile=form2.save(commit = False)
            profile.user=user
            profile.role=request.POST.get('role')
            profile.save()
            if profile.role == 'Admin':
                Group.objects.get(name='Admin').user_set.add(user)
            else:
                Group.objects.get(name='Users').user_set.add(user)
            return render(request, 'examapp2/login.html')
    else:
        form1=SignUpForm()
        form2=UserProfileInfoForm()


    return render(request, 'examapp2/signup.html', {'form1':form1, 'form2':form2})
def first_view(request):
    return render(request, 'examapp2/first.html')

def thanks_view(request):
    return render(request, 'examapp2/thanks.html')

def user_login(request):
    if request.method=='POST':
        username=request.POST.get('username')
        password=request.POST.get('password')
        user=authenticate(username=username, password=password)
        if user:
            if user.is_active:
                login(request,user)

                if user.groups.filter(name='Users').exists():
                    return render(request, 'examapp2/first.html')
                elif user.groups.filter(name='Admin').exists():
                    return render(request, 'examapp2/second.html')
                else:
                    return HttpResponse('You are not authorised to access')
            else:
                return HttpResponse('Your account is inactive')
        else:
            return HttpResponse('Someone has logged into your account')
    else:
        return render(request, 'examapp2/login.html')

sixthanswer=0
sixthquestionmarks=0
seventhanswer=0
seventhquestionmarks=0
eigthanswer=""
eigthquestionmarks=0
ninthanswer=""
ninthquestionmarks=0

def Sixthquestion(request):
    if request.method=='POST':
        answer=request.POST['sixthanswer']
        global sixthanswer
        sixthanswer=answer
        if answer=='4':
            global sixthquestionmarks
            sixthquestionmarks=2
        else:
            sixthquestionmarks=0
        return render(request,'examapp2/SEVENTHQUESTION.html')
    return render(request, 'examapp2/SIXTHQUESTION.html')

def Seventhquestion(request):
    if request.method=="POST":
        answer=request.POST['seventhanswer']
        global seventhanswer
        seventhanswer=answer
        if answer=='3':
            global seventhquestionmarks
            seventhquestionmarks=2
        else:
            seventhquestionmarks=0
        return render(request,'examapp2/EIGTHQUESTION.html')

def Eigthquestion(request):
    if request.method=='POST':
        answer=request.POST['eigthanswer']
        global eigthanswer
        eigthanswer=answer
        return render(request, 'examapp2/NINTHQUESTION.html')

def Ninthquestion(request):
    if request.method=='POST':
        answer=request.POST['ninthanswer']
        global ninthanswer
        ninthanswer=answer
        return render(request, 'examapp2/thanks.html')

def Quit(request):
    global sixthanswer
    global sixthquestionmarks
    global seventhanswer
    global seventhquestionmarks
    global eigthanswer
    global eigthquestionmarks
    global ninthanswer
    global ninthquestionmarks
    name=request.user.username
    total=TotalScore(username=name,sixthquestionanswer=sixthanswer,sixthquestionmarks=sixthquestionmarks,seventhquestionanswer=seventhanswer,seventhquestionmarks=seventhquestionmarks,eigthquestionanswer=eigthanswer,eigthquestionmarks=eigthquestionmarks,ninthquestionanswer=ninthanswer,ninthquestionmarks=ninthquestionmarks)
    total.save()
    return redirect('/user_login')

def Thanks(request):
    return render(request, 'examapp2/thanks.html')

def Dashboard(request):
    employees = TotalScore.objects.all()
    return render(request, 'examapp2/DISPLAY.html', {'employees':employees})

def Validate(request,id):
    employee = TotalScore.objects.get(id=id)
    if request.method =='POST':
        form = TotalScoreForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            return redirect('/dashboard')
    return render(request, 'examapp2/VALIDATE.html', {'employee':employee})



def TotalMarks(request):
    name=request.user.username
    marks=TotalScore.objects.get(username=name)
    m6=marks.sixthquestionmarks
    m7=marks.seventhquestionmarks
    m8=marks.eigthquestionmarks
    m9=marks.ninthquestionmarks
    total=m6+m7+m8+m9
    return render(request,'examapp2/TOTALSCORE.html',{'total':total})

def Export(request):
    page = requests.get('http://127.0.0.1:8000/dashboard/')
    soup = BeautifulSoup(page.text, 'html.parser')
    table = soup.find('table', {"class":"rambo"})
    row1 = table.find('thead')
    row2 = table.find_all("tr")
    book = openpyxl.Workbook()
    book.create_sheet("data")
    i=1
    j=1
    for h in row1.find_all('th'):
        sheet = book['data']
        sheet.cell(row=i,column=j).value = h.text
        sheet.cell(row=i,column=j).font = openpyxl.styles.Font(bold=True)
        j=j+1
    a = 1
    b = 1
    for r in row2:
        a=a+1
        b=1
        for d in r.find_all('td'):
            sheet = book['data']
            sheet.cell(row=a,column=b).value = d.text
            b = b+1

    book.save("sumanth.xlsx")
    return render(request, "examapp2/EXCEL.html")
